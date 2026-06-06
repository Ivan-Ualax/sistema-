import re
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment # type: ignore
from .permissoes import apenas_admin
from xhtml2pdf import pisa  # type: ignore
from django.template.loader import get_template
from urllib.parse import quote
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Sum, F, Q, Value
from django.db.models.functions import TruncMonth, Replace
from django.db import transaction
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4



from datetime import date
from dateutil.relativedelta import relativedelta  # type: ignore
import decimal
import json

from .models import (
    Cliente, Financeiro, Produto, Venda, ItemVenda,
    Mensagem, Compra, ParcelaVenda, Despesa, Fornecedor, NotaFiscal, VendaFiado, PagamentoFiado,CaixaPDV
)

# DASHBOARD


def dados_dashboard():
    hoje = date.today()
    meses_base = []

    for i in range(5, -1, -1):
        mes = hoje.month - i
        ano = hoje.year

        while mes <= 0:
            mes += 12
            ano -= 1

        meses_base.append((ano, mes))

    meses = []
    entradas = []
    saidas = []

    for ano, mes in meses_base:
        meses.append(f"{mes:02d}/{ano}")
        entradas.append(0)
        saidas.append(0)

    dados = Financeiro.objects.annotate(
        mes=TruncMonth('data')
    ).values('mes', 'tipo').annotate(
        total=Sum('valor')
    ).order_by('mes')

    for item in dados:
        chave = f"{item['mes'].month:02d}/{item['mes'].year}"

        if chave in meses:
            index = meses.index(chave)

            if item['tipo'] == 'entrada':
                entradas[index] = float(item['total'])
            elif item['tipo'] == 'saida':
                saidas[index] = float(item['total'])

    return meses, entradas, saidas


def inadimplencia(request):
    hoje = date.today()

    for parcela in ParcelaVenda.objects.filter(status='pendente', vencimento__lt=hoje):
        parcela.status = 'atrasada'
        parcela.save()

    parcelas = ParcelaVenda.objects.filter(
        status='atrasada',
        baixado=False
    ).select_related('venda__cliente').order_by('venda__cliente__nome', 'vencimento')

    clientes_dict = {}

    for parcela in parcelas:
        cliente = parcela.venda.cliente
        dias_atraso = (hoje - parcela.vencimento).days

        telefone_limpo = re.sub(r'\D', '', cliente.telefone or '')

        if telefone_limpo and not telefone_limpo.startswith('55'):
            telefone_limpo = '55' + telefone_limpo

        if cliente.id not in clientes_dict:
            clientes_dict[cliente.id] = {
                'cliente': cliente,
                'telefone': telefone_limpo,
                'total_devido': 0,
                'maior_atraso': 0,
                'parcelas': []
            }

        clientes_dict[cliente.id]['total_devido'] += parcela.valor

        if dias_atraso > clientes_dict[cliente.id]['maior_atraso']:
            clientes_dict[cliente.id]['maior_atraso'] = dias_atraso

        clientes_dict[cliente.id]['parcelas'].append({
            'parcela': parcela,
            'dias_atraso': dias_atraso
        })

    inadimplentes = []

    for dados in clientes_dict.values():
        cliente = dados['cliente']
        total = dados['total_devido']

        mensagem = (
            f"Olá, {cliente.nome}. Consta em nosso sistema um valor em atraso "
            f"de R$ {total:.2f}. Por favor, entre em contato para regularização."
        )

        dados['mensagem_whatsapp'] = quote(mensagem)
        inadimplentes.append(dados)

    total_atrasado = sum(item['total_devido'] for item in inadimplentes)
    total_clientes = len(inadimplentes)
    total_parcelas = parcelas.count()

    return render(request, 'inadimplencia.html', {
        'inadimplentes': inadimplentes,
        'total_atrasado': total_atrasado,
        'total_clientes': total_clientes,
        'total_parcelas': total_parcelas,
    })


def home(request):
    for parcela in ParcelaVenda.objects.filter(status='pendente', vencimento__lt=now().date()):
        parcela.atualizar_status()

    clientes_total = Cliente.objects.count()
    ativos = Cliente.objects.filter(status=True).count()
    inativos = Cliente.objects.filter(status=False).count()

    entrada = Financeiro.objects.filter(tipo='entrada').aggregate(total=Sum('valor'))['total'] or 0
    saida = Financeiro.objects.filter(tipo='saida').aggregate(total=Sum('valor'))['total'] or 0
    saldo = entrada - saida

    total_compras = Financeiro.objects.filter(
        tipo='saida',
        compra__isnull=False
    ).aggregate(total=Sum('valor'))['total'] or 0

    total_despesas = Financeiro.objects.filter(
        tipo='saida',
        despesa__isnull=False
    ).aggregate(total=Sum('valor'))['total'] or 0

    lucro_bruto = entrada - total_compras
    lucro_liquido = entrada - total_compras - total_despesas

    total_produtos = Produto.objects.count()
    estoque_baixo = Produto.objects.filter(quantidade__lte=F('estoque_minimo')).count()
    produtos_baixos = Produto.objects.filter(quantidade__lte=F('estoque_minimo'))

    valor_estoque = Produto.objects.aggregate(
        total=Sum(F('quantidade') * F('preco_custo'))
    )['total'] or 0

    meses, entradas_mes, saidas_mes = dados_dashboard()

    top_produtos = ItemVenda.objects.values('produto__nome').annotate(
        total=Sum('quantidade')
    ).order_by('-total')[:5]

    top_clientes = Venda.objects.values('cliente__nome').annotate(
        total=Sum('total_com_juros')
    ).order_by('-total')[:5]

    top_compras = Compra.objects.values('produto__nome').annotate(
        total=Sum('quantidade')
    ).order_by('-total')[:5]

    parcelas_pendentes = ParcelaVenda.objects.filter(status='pendente').count()

    valor_a_receber = ParcelaVenda.objects.filter(status='pendente').aggregate(
        total=Sum('valor')
    )['total'] or 0

    parcelas_vencidas = ParcelaVenda.objects.filter(status='atrasada').count()

    inadimplentes = ParcelaVenda.objects.filter(
        status='atrasada'
    ).values('venda__cliente__nome').distinct().count()

    return render(request, 'home.html', {
        'clientes_total': clientes_total,
        'ativos': ativos,
        'inativos': inativos,
        'entrada': entrada,
        'saida': saida,
        'saldo': saldo,
        'total_compras': total_compras,
        'total_despesas': total_despesas,
        'lucro_bruto': lucro_bruto,
        'lucro_liquido': lucro_liquido,
        'total_produtos': total_produtos,
        'estoque_baixo': estoque_baixo,
        'produtos_baixos': produtos_baixos,
        'valor_estoque': valor_estoque,
        'parcelas_pendentes': parcelas_pendentes,
        'valor_a_receber': valor_a_receber,
        'parcelas_vencidas': parcelas_vencidas,
        'inadimplentes': inadimplentes,
        'meses': json.dumps(meses),
        'entradas_mes': json.dumps(entradas_mes),
        'saidas_mes': json.dumps(saidas_mes),
        'top_produtos': top_produtos,
        'top_clientes': top_clientes,
        'top_compras': top_compras,
    })



# LOGIN


def login_usuario(request):
    if request.method == 'POST':
        usuario = request.POST.get('usuario')
        senha = request.POST.get('senha')

        user = authenticate(request, username=usuario, password=senha)

        if user is not None:
            login(request, user)
            return redirect('/')
        else:
            return render(request, 'login.html', {
                'erro': 'Usuário ou senha inválidos'
            })

    return render(request, 'login.html')


def logout_usuario(request):
    logout(request)
    return redirect('/login/')



# CLIENTES

def lista_clientes(request):
    q = request.GET.get('q', '')
    clientes = Cliente.objects.all()

    if q:
        busca = (
            q.replace('.', '')
             .replace('-', '')
             .replace('/', '')
             .replace(' ', '')
        )

        clientes = clientes.annotate(
            cpf_limpo=Replace(
                Replace(
                    Replace(
                        Replace('cpf', Value('.'), Value('')),
                        Value('-'), Value('')
                    ),
                    Value('/'), Value('')
                ),
                Value(' '), Value('')
            )
        ).filter(
            Q(nome__icontains=q) |
            Q(cpf_limpo__icontains=busca)
        )

    return render(request, 'lista_clientes.html', {
        'clientes': clientes,
        'q': q
    })


def cadastrar(request):
    if request.method == 'POST':
        Cliente.objects.create(
            nome=request.POST.get('nome'),
            cpf=request.POST.get('cpf') or None,
            telefone=request.POST.get('telefone') or None,
            localizacao=request.POST.get('localizacao') or None,
            status=request.POST.get('status') == 'True'
        )
        return redirect('/clientes/')

    return render(request, 'cadastrar.html')


def editar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)

    if request.method == 'POST':
        cpf = request.POST.get('cpf')

        cpf_existe = Cliente.objects.filter(cpf=cpf).exclude(id=cliente.id).exists()

        if cpf_existe:
            return render(request, 'editar_cliente.html', {
                'cliente': cliente,
                'erro': 'Já existe outro cliente cadastrado com este CPF.'
            })

        cliente.nome = request.POST.get('nome')
        cliente.cpf = cpf
        cliente.telefone = request.POST.get('telefone')
        cliente.localizacao = request.POST.get('localizacao')
        cliente.status = request.POST.get('status') == 'True'
        cliente.save()

        return redirect('/clientes/')

    return render(request, 'editar_cliente.html', {
        'cliente': cliente
    })


@apenas_admin
def deletar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)
    cliente.delete()
    return redirect('/clientes/')


# FINANCEIRO


def financeiro(request):
    registros = Financeiro.objects.all().order_by('-id')

    entrada = Financeiro.objects.filter(tipo='entrada').aggregate(total=Sum('valor'))['total'] or 0
    saida = Financeiro.objects.filter(tipo='saida').aggregate(total=Sum('valor'))['total'] or 0
    saldo = entrada - saida

    return render(request, 'financeiro.html', {
        'registros': registros,
        'entrada': entrada,
        'saida': saida,
        'saldo': saldo,
    })


@apenas_admin
def editar_financeiro(request, id):
    registro = get_object_or_404(Financeiro, id=id)

    if request.method == 'POST':
        registro.descricao = request.POST.get('descricao')
        registro.valor = decimal.Decimal(request.POST.get('valor', '0').replace(',', '.'))
        registro.tipo = request.POST.get('tipo')
        registro.save()
        return redirect('/financeiro/')

    return render(request, 'editar_financeiro.html', {'registro': registro})


@apenas_admin
def deletar_financeiro(request, id):
    registro = get_object_or_404(Financeiro, id=id)
    registro.delete()
    return redirect('/financeiro/')



# ESTOQUE

def estoque(request):
    q = request.GET.get('q')
    filtro = request.GET.get('filtro')
    ordenar = request.GET.get('ordenar')

    produtos = Produto.objects.all()

    if q:
        produtos = produtos.filter(nome__icontains=q)

    if filtro == 'baixo':
        produtos = produtos.filter(quantidade__lte=F('estoque_minimo'))
    elif filtro == 'sem':
        produtos = produtos.filter(quantidade__lte=0)

    if ordenar == 'nome':
        produtos = produtos.order_by('nome')
    elif ordenar == 'quantidade':
        produtos = produtos.order_by('-quantidade')
    elif ordenar == 'preco':
        produtos = produtos.order_by('-preco_venda')

    return render(request, 'estoque.html', {
        'produtos': produtos,
        'q': q,
        'filtro': filtro,
        'ordenar': ordenar
    })


def adicionar_produto(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')

        quantidade = decimal.Decimal(
            request.POST.get('quantidade', '0').replace(',', '.')
        )

        preco_custo = request.POST.get('preco_custo', '').strip()
        preco_venda = request.POST.get('preco_venda', '').strip()

        estoque_minimo = decimal.Decimal(
            request.POST.get('estoque_minimo', '5').replace(',', '.')
        )

        if preco_custo == '' or preco_venda == '':
            return HttpResponse("Erro: preencha preço de custo e venda")

        Produto.objects.create(
            nome=nome,
            quantidade=quantidade,
            preco_custo=decimal.Decimal(preco_custo.replace(',', '.')),
            preco_venda=decimal.Decimal(preco_venda.replace(',', '.')),
            estoque_minimo=estoque_minimo
        )

        return redirect('/estoque/')

    return render(request, 'add_estoque.html')


def editar_produto(request, id):
    produto = get_object_or_404(Produto, id=id)

    if request.method == 'POST':
        produto.nome = request.POST.get('nome')

        produto.quantidade = decimal.Decimal(
            request.POST.get('quantidade', '0').replace(',', '.')
        )

        produto.preco_custo = decimal.Decimal(
            request.POST.get('preco_custo', '0').replace(',', '.')
        )

        produto.preco_venda = decimal.Decimal(
            request.POST.get('preco_venda', '0').replace(',', '.')
        )

        produto.estoque_minimo = decimal.Decimal(
            request.POST.get('estoque_minimo', '5').replace(',', '.')
        )

        produto.save()
        return redirect('/estoque/')

    return render(request, 'editar_produto.html', {'produto': produto})


@apenas_admin
def deletar_produto(request, id):
    produto = get_object_or_404(Produto, id=id)

    tem_venda = ItemVenda.objects.filter(produto=produto).exists()

    if tem_venda:
        return render(request, 'erro_exclusao.html', {
            'mensagem': 'Não é possível apagar este produto.',
            'detalhe': 'Este produto possui vendas vinculadas. Para manter o histórico correto, ele não pode ser excluído.'
        })

    with transaction.atomic():
        Financeiro.objects.filter(compra__produto=produto).delete()
        Compra.objects.filter(produto=produto).delete()
        produto.delete()

    return redirect('/estoque/')


# COMPRAS

def compras(request):
    return render(request, 'compras.html', {
        'compras': Compra.objects.all().order_by('-id')
    })


def adicionar_compra(request):
    fornecedores = Fornecedor.objects.filter(ativo=True).order_by('nome')

    if request.method == 'POST':
        nome_produto = request.POST.get('nome')
        fornecedor_id = request.POST.get('fornecedor')

        quantidade = decimal.Decimal(
            request.POST.get('quantidade', '0').replace(',', '.')
        )

        fornecedor_obj = None
        if fornecedor_id:
            fornecedor_obj = Fornecedor.objects.filter(id=fornecedor_id).first()

        fornecedor_nome = fornecedor_obj.nome if fornecedor_obj else 'Não informado'

        preco_custo = decimal.Decimal(
            request.POST.get('preco_custo', '0').replace(',', '.')
        )

        preco_venda = decimal.Decimal(
            request.POST.get('preco_venda', '0').replace(',', '.')
        )

        estoque_minimo = decimal.Decimal(
            request.POST.get('estoque_minimo', '5').replace(',', '.')
        )

        if not nome_produto:
            return HttpResponse("Nome do produto obrigatório")

        if quantidade <= 0:
            return HttpResponse("Quantidade inválida")

        if preco_custo <= 0 or preco_venda <= 0:
            return HttpResponse("Preço de custo e venda devem ser maiores que zero")

        with transaction.atomic():
            produto, criado = Produto.objects.get_or_create(
                nome=nome_produto,
                defaults={
                    'quantidade': 0,
                    'preco_custo': preco_custo,
                    'preco_venda': preco_venda,
                    'estoque_minimo': estoque_minimo,
                    'data_ultima_compra': now(),
                    'fornecedor': fornecedor_nome,
                    'fornecedor_cadastro': fornecedor_obj,
                }
            )

            produto.quantidade += quantidade
            produto.preco_custo = preco_custo
            produto.preco_venda = preco_venda
            produto.estoque_minimo = estoque_minimo
            produto.data_ultima_compra = now()
            produto.fornecedor = fornecedor_nome
            produto.fornecedor_cadastro = fornecedor_obj
            produto.save()

            compra = Compra.objects.create(
                produto=produto,
                fornecedor=fornecedor_nome,
                fornecedor_cadastro=fornecedor_obj,
                quantidade=quantidade,
                preco=preco_custo,
                preco_venda=preco_venda
            )

            Financeiro.objects.create(
                descricao=f"Compra de {produto.nome}",
                valor=preco_custo * quantidade,
                tipo='saida',
                compra=compra
            )

        return redirect('/estoque/')

    return render(request, 'add_compra.html', {
        'fornecedores': fornecedores
    })


@apenas_admin
def editar_compra(request, id):
    compra = get_object_or_404(Compra, id=id)
    fornecedores = Fornecedor.objects.filter(ativo=True).order_by('nome')

    if request.method == 'POST':
        fornecedor_id = request.POST.get('fornecedor')

        fornecedor_obj = None
        if fornecedor_id:
            fornecedor_obj = Fornecedor.objects.filter(id=fornecedor_id).first()

        fornecedor_nome = fornecedor_obj.nome if fornecedor_obj else 'Não informado'

        quantidade = decimal.Decimal(
            request.POST.get('quantidade', '0').replace(',', '.')
        )

        preco = decimal.Decimal(
            request.POST.get('preco', '0').replace(',', '.')
        )

        preco_venda = decimal.Decimal(
            request.POST.get('preco_venda', '0').replace(',', '.')
        )

        with transaction.atomic():
            produto = compra.produto

            produto.quantidade -= compra.quantidade
            produto.quantidade += quantidade
            produto.preco_custo = preco
            produto.preco_venda = preco_venda
            produto.fornecedor = fornecedor_nome
            produto.fornecedor_cadastro = fornecedor_obj
            produto.save()

            compra.fornecedor = fornecedor_nome
            compra.fornecedor_cadastro = fornecedor_obj
            compra.quantidade = quantidade
            compra.preco = preco
            compra.preco_venda = preco_venda
            compra.save()

            financeiro = Financeiro.objects.filter(compra=compra).first()
            if financeiro:
                financeiro.valor = preco * quantidade
                financeiro.descricao = f"Compra editada de {produto.nome}"
                financeiro.save()

        return redirect('/compras/')

    return render(request, 'editar_compra.html', {
        'compra': compra,
        'fornecedores': fornecedores
    })


@apenas_admin
def deletar_compra(request, id):
    compra = get_object_or_404(Compra, id=id)

    with transaction.atomic():
        produto = compra.produto
        produto.quantidade -= compra.quantidade
        produto.save()

        Financeiro.objects.filter(compra=compra).delete()
        compra.delete()

    return redirect('/compras/')

# VENDAS

def vendas(request):
    return render(request, 'vendas.html', {
        'vendas': Venda.objects.all().order_by('-id')
    })


def criar_venda(request):
    if request.method == 'POST':
        cliente = get_object_or_404(Cliente, id=request.POST.get('cliente'))
        produto_id = request.POST.get('produto')

        quantidade = decimal.Decimal(
            request.POST.get('quantidade', '0').replace(',', '.')
        )

        parcelado = request.POST.get('parcelado') == 'on'
        quantidade_parcelas = int(request.POST.get('quantidade_parcelas') or 1)

        juros_percentual = decimal.Decimal(
            (request.POST.get('juros_percentual') or '0').replace(',', '.')
        )

        if quantidade <= 0:
            return HttpResponse("Quantidade inválida")

        with transaction.atomic():
            produto = Produto.objects.select_for_update().get(id=produto_id)

            if produto.quantidade < quantidade:
                return HttpResponse("Estoque insuficiente")

            total = produto.preco_venda * quantidade

            if parcelado:
                taxa = juros_percentual / 100
                total_com_juros = total + (total * taxa * quantidade_parcelas)
                valor_parcela = total_com_juros / quantidade_parcelas
            else:
                total_com_juros = total
                valor_parcela = total
                quantidade_parcelas = 1
                juros_percentual = 0

            venda = Venda.objects.create(
                cliente=cliente,
                total=total,
                total_com_juros=total_com_juros,
                parcelado=parcelado,
                quantidade_parcelas=quantidade_parcelas,
                juros_percentual=juros_percentual
            )

            lucro = (produto.preco_venda - produto.preco_custo) * quantidade

            ItemVenda.objects.create(
                venda=venda,
                produto=produto,
                quantidade=quantidade,
                preco=produto.preco_venda,
                preco_custo=produto.preco_custo,
                lucro=lucro
            )

            produto.quantidade -= quantidade
            produto.save()

            if parcelado:
                for i in range(1, quantidade_parcelas + 1):
                    ParcelaVenda.objects.create(
                        venda=venda,
                        numero=i,
                        valor=valor_parcela,
                        vencimento=date.today() + relativedelta(months=i),
                        status='pendente',
                        baixado=False
                    )
            else:
                Financeiro.objects.create(
                    descricao=f"Venda à vista para {cliente.nome}",
                    valor=total_com_juros,
                    tipo='entrada',
                    venda=venda
                )

        return redirect('/vendas/')

    return render(request, 'criar_venda.html', {
        'clientes': Cliente.objects.all(),
        'produtos': Produto.objects.all()
    })


@apenas_admin
def deletar_venda(request, id):
    venda = get_object_or_404(Venda, id=id)

    with transaction.atomic():
        itens = ItemVenda.objects.filter(venda=venda)

        for item in itens:
            produto = item.produto
            produto.quantidade += item.quantidade
            produto.save()

        Financeiro.objects.filter(venda=venda).delete()
        Financeiro.objects.filter(parcela__venda=venda).delete()
        ParcelaVenda.objects.filter(venda=venda).delete()
        ItemVenda.objects.filter(venda=venda).delete()
        venda.delete()

    return redirect('/vendas/')


# PARCELAS


def lista_parcelas(request):
    vendas = Venda.objects.filter(parcelado=True).select_related('cliente').order_by('-id')

    return render(request, 'parcelas.html', {
        'vendas': vendas
    })


def detalhe_parcelas(request, id):
    venda = get_object_or_404(Venda, id=id, parcelado=True)

    for parcela in ParcelaVenda.objects.filter(venda=venda, status='pendente'):
        parcela.atualizar_status()

    parcelas = ParcelaVenda.objects.filter(venda=venda).order_by('numero')

    total_pago = parcelas.filter(status='paga').aggregate(total=Sum('valor'))['total'] or 0
    total_pendente = parcelas.filter(status='pendente').aggregate(total=Sum('valor'))['total'] or 0
    total_baixado = parcelas.filter(baixado=True).aggregate(total=Sum('valor'))['total'] or 0

    return render(request, 'detalhe_parcelas.html', {
        'venda': venda,
        'parcelas': parcelas,
        'total_pago': total_pago,
        'total_pendente': total_pendente,
        'total_baixado': total_baixado,
        'today': now().date()
    })


@apenas_admin
def editar_parcela(request, id):
    parcela = get_object_or_404(ParcelaVenda, id=id)

    if parcela.baixado:
        return render(request, 'erro_exclusao.html', {
            'mensagem': 'Esta parcela já foi baixada.',
            'detalhe': 'Não é recomendado editar valor ou vencimento de uma parcela que já entrou no financeiro.'
        })

    if request.method == 'POST':
        valor = decimal.Decimal(
            request.POST.get('valor', '0').replace(',', '.')
        )
        vencimento = request.POST.get('vencimento')
        status = request.POST.get('status')

        parcela.valor = valor
        parcela.vencimento = vencimento
        parcela.status = status

        if status == 'paga':
            parcela.data_pagamento = now().date()
        else:
            parcela.data_pagamento = None

        parcela.save()

        return redirect(f'/parcelas/venda/{parcela.venda.id}/')

    return render(request, 'editar_parcela.html', {
        'parcela': parcela
    })


def marcar_pago(request, id):
    parcela = get_object_or_404(ParcelaVenda, id=id)

    if parcela.status in ['pendente', 'atrasada']:
        parcela.status = 'paga'
        parcela.data_pagamento = now().date()
        parcela.save()

    return redirect(f'/parcelas/venda/{parcela.venda.id}/')


def baixar_parcela(request, id):
    parcela = get_object_or_404(ParcelaVenda, id=id)

    if parcela.status == 'paga' and not parcela.baixado:
        Financeiro.objects.create(
            descricao=f"Baixa parcela {parcela.numero} - {parcela.venda.cliente.nome}",
            valor=parcela.valor,
            tipo='entrada',
            parcela=parcela
        )

        parcela.baixado = True
        parcela.save()

    return redirect(f'/parcelas/venda/{parcela.venda.id}/')


def imprimir_parcela(request, id):
    parcela = get_object_or_404(ParcelaVenda, id=id)

    return render(request, 'comprovante.html', {
        'parcela': parcela
    })


@apenas_admin
def deletar_venda_parcelada(request, id):
    venda = get_object_or_404(Venda, id=id, parcelado=True)

    parcelas_pendentes = ParcelaVenda.objects.filter(
        venda=venda,
        status__in=['pendente', 'atrasada']
    ).exists()

    if parcelas_pendentes:
        return render(request, 'erro_exclusao.html', {
            'mensagem': 'Não é possível apagar esta venda parcelada.',
            'detalhe': 'Ainda existem parcelas pendentes ou atrasadas. Quite todas as parcelas antes de excluir a venda.'
        })

    with transaction.atomic():
        Financeiro.objects.filter(parcela__venda=venda).delete()
        Financeiro.objects.filter(venda=venda).delete()
        ParcelaVenda.objects.filter(venda=venda).delete()
        ItemVenda.objects.filter(venda=venda).delete()
        venda.delete()

    return redirect('/parcelas/')



# MENSAGENS


def enviar_mensagem(request):
    if request.method == 'POST':
        Mensagem.objects.create(
            nome=request.POST.get('nome'),
            email=request.POST.get('email'),
            mensagem=request.POST.get('mensagem')
        )

    return redirect('/')



# RELATÓRIO


def relatorio_lucro(request):
    receita = Financeiro.objects.filter(tipo='entrada').aggregate(total=Sum('valor'))['total'] or 0
    custo = Financeiro.objects.filter(tipo='saida').aggregate(total=Sum('valor'))['total'] or 0

    return render(request, 'relatorio.html', {
        'receita': receita,
        'custo': custo,
        'lucro': receita - custo
    })



# DESPESAS


def despesas(request):
    despesas = Despesa.objects.all().order_by('-id')

    total_despesas = Despesa.objects.aggregate(
        total=Sum('valor')
    )['total'] or 0

    return render(request, 'despesas.html', {
        'despesas': despesas,
        'total_despesas': total_despesas
    })


def adicionar_despesa(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        valor = decimal.Decimal(
            request.POST.get('valor', '0').replace(',', '.')
        )
        categoria = request.POST.get('categoria')
        observacao = request.POST.get('observacao')

        with transaction.atomic():
            despesa = Despesa.objects.create(
                nome=nome,
                valor=valor,
                categoria=categoria,
                observacao=observacao
            )

            Financeiro.objects.create(
                descricao=f"Despesa: {nome}",
                valor=valor,
                tipo='saida',
                despesa=despesa
            )

        return redirect('/despesas/')

    return render(request, 'add_despesa.html')


@apenas_admin
def deletar_despesa(request, id):
    despesa = get_object_or_404(Despesa, id=id)

    with transaction.atomic():
        Financeiro.objects.filter(despesa=despesa).delete()
        despesa.delete()

    return redirect('/despesas/')



# PDF RELATÓRIOS


def gerar_pdf(template_src, context_dict, nome_arquivo):
    template = get_template(template_src)
    html = template.render(context_dict)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{nome_arquivo}"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Erro ao gerar PDF")

    return response


def relatorio_vendas_pdf(request):
    vendas = Venda.objects.all().select_related('cliente').order_by('-data')

    total_vendas = vendas.aggregate(
        total=Sum('total_com_juros')
    )['total'] or 0

    return gerar_pdf('pdf_vendas.html', {
        'vendas': vendas,
        'total_vendas': total_vendas,
        'data_geracao': now()
    }, 'relatorio_vendas.pdf')


def relatorio_financeiro_pdf(request):
    registros = Financeiro.objects.all().order_by('-data')

    entradas = Financeiro.objects.filter(tipo='entrada').aggregate(
        total=Sum('valor')
    )['total'] or 0

    saidas = Financeiro.objects.filter(tipo='saida').aggregate(
        total=Sum('valor')
    )['total'] or 0

    saldo = entradas - saidas

    return gerar_pdf('pdf_financeiro.html', {
        'registros': registros,
        'entradas': entradas,
        'saidas': saidas,
        'saldo': saldo,
        'data_geracao': now()
    }, 'relatorio_financeiro.pdf')


def relatorio_inadimplencia_pdf(request):
    hoje = date.today()

    for parcela in ParcelaVenda.objects.filter(status='pendente', vencimento__lt=hoje):
        parcela.status = 'atrasada'
        parcela.save()

    parcelas = ParcelaVenda.objects.filter(
        status='atrasada',
        baixado=False
    ).select_related('venda__cliente').order_by('venda__cliente__nome', 'vencimento')

    total_atrasado = parcelas.aggregate(
        total=Sum('valor')
    )['total'] or 0

    return gerar_pdf('pdf_inadimplencia.html', {
        'parcelas': parcelas,
        'total_atrasado': total_atrasado,
        'total_parcelas': parcelas.count(),
        'data_geracao': now()
    }, 'relatorio_inadimplencia.pdf')


# FORNECEDORES


def fornecedores(request):
    q = request.GET.get('q', '')

    lista = Fornecedor.objects.all().order_by('nome')

    if q:
        lista = lista.filter(
            Q(nome__icontains=q) |
            Q(cnpj__icontains=q) |
            Q(telefone__icontains=q)
        )

    return render(request, 'fornecedores.html', {
        'fornecedores': lista,
        'q': q
    })


def adicionar_fornecedor(request):
    if request.method == 'POST':
        Fornecedor.objects.create(
            nome=request.POST.get('nome'),
            cnpj=request.POST.get('cnpj'),
            telefone=request.POST.get('telefone'),
            email=request.POST.get('email'),
            endereco=request.POST.get('endereco'),
            observacao=request.POST.get('observacao'),
            ativo=request.POST.get('ativo') == 'True'
        )

        return redirect('/fornecedores/')

    return render(request, 'add_fornecedor.html')


def editar_fornecedor(request, id):
    fornecedor = get_object_or_404(Fornecedor, id=id)

    if request.method == 'POST':
        fornecedor.nome = request.POST.get('nome')
        fornecedor.cnpj = request.POST.get('cnpj')
        fornecedor.telefone = request.POST.get('telefone')
        fornecedor.email = request.POST.get('email')
        fornecedor.endereco = request.POST.get('endereco')
        fornecedor.observacao = request.POST.get('observacao')
        fornecedor.ativo = request.POST.get('ativo') == 'True'
        fornecedor.save()

        return redirect('/fornecedores/')

    return render(request, 'editar_fornecedor.html', {
        'fornecedor': fornecedor
    })


@apenas_admin
def deletar_fornecedor(request, id):
    fornecedor = get_object_or_404(Fornecedor, id=id)
    fornecedor.delete()

    return redirect('/fornecedores/')
def relatorio_financeiro_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Financeiro"

    ws.append(["Descrição", "Tipo", "Valor", "Data"])

    for registro in Financeiro.objects.all().order_by("-data"):
        ws.append([
            registro.descricao,
            registro.tipo,
            float(registro.valor),
            registro.data.strftime("%d/%m/%Y")
        ])

    # Cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center")

    # Largura das colunas
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="relatorio_financeiro.xlsx"'

    wb.save(response)

    return response

# NOTAS FISCAIS


def notas_fiscais(request):
    notas = NotaFiscal.objects.all().select_related('cliente', 'venda').order_by('-id')

    return render(request, 'notas_fiscais.html', {
        'notas': notas
    })


def adicionar_nota_fiscal(request):
    clientes = Cliente.objects.filter(status=True).order_by('nome')
    vendas = Venda.objects.all().select_related('cliente').order_by('-id')

    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        venda_id = request.POST.get('venda')

        cliente = Cliente.objects.filter(id=cliente_id).first() if cliente_id else None
        venda = Venda.objects.filter(id=venda_id).first() if venda_id else None

        NotaFiscal.objects.create(
            numero=request.POST.get('numero'),
            chave_acesso=request.POST.get('chave_acesso'),
            tipo=request.POST.get('tipo'),
            status=request.POST.get('status'),
            cliente=cliente,
            venda=venda,
            valor_total=request.POST.get('valor_total') or 0,
            observacao=request.POST.get('observacao')
        )

        return redirect('/notas-fiscais/')

    return render(request, 'add_nota_fiscal.html', {
        'clientes': clientes,
        'vendas': vendas
    })


@apenas_admin
def deletar_nota_fiscal(request, id):
    nota = get_object_or_404(NotaFiscal, id=id)
    nota.delete()

    return redirect('/notas-fiscais/')


# 🤝 VENDA FIADO

def vendas_fiado(request):
    fiados = VendaFiado.objects.select_related('cliente').order_by('-id')

    total_em_aberto = sum(
        f.saldo_devedor() for f in fiados if f.status == 'aberta'
    )

    return render(request, 'vendas_fiado.html', {
        'fiados': fiados,
        'total_em_aberto': total_em_aberto
    })


def criar_venda_fiado(request):
    if request.method == 'POST':
        cliente = get_object_or_404(Cliente, id=request.POST.get('cliente'))
        produto = get_object_or_404(Produto, id=request.POST.get('produto'))

        quantidade = decimal.Decimal(
            request.POST.get('quantidade', '0').replace(',', '.')
        )

        observacao = request.POST.get('observacao')

        if quantidade <= 0:
            return HttpResponse("Quantidade inválida")

        with transaction.atomic():
            produto = Produto.objects.select_for_update().get(id=produto.id)

            if produto.quantidade < quantidade:
                return HttpResponse("Estoque insuficiente")

            total = produto.preco_venda * quantidade
            total = total.quantize(decimal.Decimal('0.01'))

            lucro = (produto.preco_venda - produto.preco_custo) * quantidade
            lucro = lucro.quantize(decimal.Decimal('0.01'))

            venda = Venda.objects.create(
                cliente=cliente,
                total=total,
                total_com_juros=total,
                parcelado=False,
                quantidade_parcelas=1,
                juros_percentual=0,
                forma_pagamento='fiado'
            )
            ItemVenda.objects.create(
                venda=venda,
                produto=produto,
                quantidade=quantidade,
                preco=produto.preco_venda,
                preco_custo=produto.preco_custo,
                lucro=lucro
            )

            produto.quantidade -= quantidade
            produto.save()

            VendaFiado.objects.create(
                cliente=cliente,
                venda=venda,
                valor_total=total,
                valor_pago=0,
                status='aberta',
                observacao=observacao
            )

        return redirect('/venda-fiado/')

    return render(request, 'criar_venda_fiado.html', {
        'clientes': Cliente.objects.filter(status=True).order_by('nome'),
        'produtos': Produto.objects.all().order_by('nome')
    })


def detalhe_venda_fiado(request, id):
    fiado = get_object_or_404(VendaFiado, id=id)

    pagamentos = PagamentoFiado.objects.filter(
        venda_fiado=fiado
    ).order_by('-id')

    return render(request, 'detalhe_venda_fiado.html', {
        'fiado': fiado,
        'pagamentos': pagamentos
    })


def abater_venda_fiado(request, id):
    fiado = get_object_or_404(VendaFiado, id=id)

    if fiado.status == 'quitada':
        return redirect(f'/venda-fiado/{fiado.id}/')

    if request.method == 'POST':
        valor = decimal.Decimal(
            request.POST.get('valor', '0').replace(',', '.')
        )

        observacao = request.POST.get('observacao')

        if valor <= 0:
            return HttpResponse("Valor inválido")

        if valor > fiado.saldo_devedor():
            return HttpResponse("O valor do abatimento é maior que a dívida.")

        with transaction.atomic():
            PagamentoFiado.objects.create(
                venda_fiado=fiado,
                valor=valor,
                observacao=observacao
            )

            fiado.valor_pago += valor

            if fiado.valor_pago >= fiado.valor_total:
                fiado.status = 'quitada'

            fiado.save()

            Financeiro.objects.create(
                descricao=f"Abatimento fiado - {fiado.cliente.nome}",
                valor=valor,
                tipo='entrada',
                venda=fiado.venda
            )

        return redirect(f'/venda-fiado/{fiado.id}/')

    return render(request, 'abater_venda_fiado.html', {
        'fiado': fiado
    })


@apenas_admin
def excluir_venda_fiado(request, id):
    fiado = get_object_or_404(VendaFiado, id=id)

    if fiado.status != 'quitada':
        return redirect('/venda-fiado/')

    fiado.delete()

    return redirect('/venda-fiado/')

# PDV

@login_required
def pdv(request):
    caixa_aberto = CaixaPDV.objects.filter(status='aberto').first()

    if request.method == 'POST':

        if not caixa_aberto:
            return render(request, 'caixa_fechado.html')

        cliente_id = request.POST.get('cliente')
        cliente_rapido = request.POST.get('cliente_rapido', '').strip()
        forma_pagamento = request.POST.get('forma_pagamento')
        carrinho_json = request.POST.get('carrinho', '[]')

        try:
            carrinho = json.loads(carrinho_json)
        except:
            return HttpResponse("Carrinho inválido")

        if not carrinho:
            return HttpResponse("Carrinho vazio")

        with transaction.atomic():

            if cliente_rapido:
                cliente = Cliente.objects.create(
                    nome=cliente_rapido,
                    cpf=None,
                    telefone=None,
                    localizacao=None,
                    status=True
                )
            elif cliente_id:
                cliente = get_object_or_404(Cliente, id=cliente_id)
            else:
                cliente, criado = Cliente.objects.get_or_create(
                    nome='Consumidor Final',
                    defaults={
                        'cpf': None,
                        'telefone': None,
                        'localizacao': None,
                        'status': True
                    }
                )

            venda = Venda.objects.create(
                cliente=cliente,
                total=0,
                total_com_juros=0,
                parcelado=False,
                quantidade_parcelas=1,
                juros_percentual=0,
                forma_pagamento=forma_pagamento
            )

            total_venda = decimal.Decimal('0.00')

            for item in carrinho:
                produto_id = item.get('produto_id')

                quantidade = decimal.Decimal(
                    str(item.get('quantidade', '0')).replace(',', '.')
                )

                if quantidade <= 0:
                    return HttpResponse("Quantidade inválida")

                produto = Produto.objects.select_for_update().get(id=produto_id)

                if produto.quantidade < quantidade:
                    return HttpResponse(
                        f"Estoque insuficiente para {produto.nome}"
                    )

                subtotal = produto.preco_venda * quantidade
                subtotal = subtotal.quantize(decimal.Decimal('0.01'))

                lucro = (produto.preco_venda - produto.preco_custo) * quantidade
                lucro = lucro.quantize(decimal.Decimal('0.01'))

                ItemVenda.objects.create(
                    venda=venda,
                    produto=produto,
                    quantidade=quantidade,
                    preco=produto.preco_venda,
                    preco_custo=produto.preco_custo,
                    lucro=lucro
                )

                produto.quantidade -= quantidade
                produto.save()

                total_venda += subtotal

            total_venda = total_venda.quantize(decimal.Decimal('0.01'))

            venda.total = total_venda
            venda.total_com_juros = total_venda
            venda.save()

            if forma_pagamento == 'fiado':
                VendaFiado.objects.create(
                    cliente=cliente,
                    venda=venda,
                    valor_total=total_venda,
                    valor_pago=0,
                    status='aberta',
                    observacao='Venda realizada pelo PDV'
                )
            else:
                Financeiro.objects.create(
                    descricao=f"Venda PDV - {forma_pagamento} - {cliente.nome}",
                    valor=total_venda,
                    tipo='entrada',
                    venda=venda
                )

        return redirect('/pdv/')

    produtos = Produto.objects.all().order_by('nome')
    clientes = Cliente.objects.filter(status=True).order_by('nome')

    return render(request, 'pdv.html', {
        'produtos': produtos,
        'clientes': clientes,
        'caixa_aberto': caixa_aberto,
    })


@login_required
def pdv_vendas(request):
    vendas = Venda.objects.select_related(
        'cliente'
    ).prefetch_related(
        'itemvenda_set__produto'
    ).order_by('-id')

    return render(request, 'pdv_vendas.html', {
        'vendas': vendas
    })


@login_required
def abrir_caixa(request):
    caixa_aberto = CaixaPDV.objects.filter(status='aberto').first()

    if caixa_aberto:
        return redirect('/pdv/')

    if request.method == 'POST':
        valor_abertura = decimal.Decimal(
            request.POST.get('valor_abertura', '0').replace(',', '.')
        )

        CaixaPDV.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            valor_abertura=valor_abertura,
            status='aberto'
        )

        return redirect('/pdv/')

    return render(request, 'abrir_caixa.html')


@login_required
def fechar_caixa(request):
    caixa = CaixaPDV.objects.filter(status='aberto').first()

    if not caixa:
        return redirect('/pdv/')

    hoje = now().date()

    entradas = Financeiro.objects.filter(
        tipo='entrada',
        data=hoje
    ).aggregate(total=Sum('valor'))['total'] or 0

    operador = caixa.usuario.username if caixa.usuario else request.user.username

    if request.method == 'POST':
        caixa.valor_fechamento = entradas
        caixa.data_fechamento = now()
        caixa.status = 'fechado'
        caixa.observacao = request.POST.get('observacao')
        caixa.save()

        messages.success(request, "Caixa fechado com sucesso.")
        return redirect('/pdv/')

    return render(request, 'fechar_caixa.html', {
        'caixa': caixa,
        'data_hoje': hoje,
        'entradas': entradas,
        'operador': operador,
    })


@login_required
def pdf_caixa(request):
    caixa = CaixaPDV.objects.filter(
        status='fechado'
    ).order_by('-id').first()

    if not caixa:
        return HttpResponse(
            "Nenhum caixa fechado encontrado."
        )

    data_caixa = caixa.data_fechamento.date()

    vendas = Venda.objects.filter(
        data__date=data_caixa
    ).select_related(
        'cliente'
    ).order_by('id')

    total_dinheiro = vendas.filter(
        forma_pagamento='dinheiro'
    ).aggregate(
        total=Sum('total')
    )['total'] or 0

    total_pix = vendas.filter(
        forma_pagamento='pix'
    ).aggregate(
        total=Sum('total')
    )['total'] or 0

    total_cartao = vendas.filter(
        forma_pagamento='cartao'
    ).aggregate(
        total=Sum('total')
    )['total'] or 0

    total_fiado = vendas.filter(
        forma_pagamento='fiado'
    ).aggregate(
        total=Sum('total')
    )['total'] or 0

    total_geral = vendas.aggregate(
        total=Sum('total')
    )['total'] or 0

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        f'attachment; filename="caixa_{caixa.id}.pdf"'
    )

    pdf = canvas.Canvas(
        response,
        pagesize=A4
    )

    y = 800

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(
        180,
        y,
        "FECHAMENTO DE CAIXA"
    )

    y -= 40

    operador = (
        caixa.usuario.username
        if caixa.usuario
        else "Não informado"
    )

    pdf.setFont("Helvetica", 11)

    pdf.drawString(
        50,
        y,
        f"Data: {caixa.data_fechamento.strftime('%d/%m/%Y')}"
    )

    y -= 20

    pdf.drawString(
        50,
        y,
        f"Operador: {operador}"
    )

    y -= 20

    pdf.drawString(
        50,
        y,
        f"Abertura: R$ {caixa.valor_abertura:.2f}"
    )

    y -= 20

    pdf.drawString(
        50,
        y,
        f"Fechamento: R$ {caixa.valor_fechamento:.2f}"
    )

    y -= 20

    pdf.drawString(
        50,
        y,
        f"Observação: {caixa.observacao or 'Nenhuma'}"
    )

    y -= 35

    pdf.setFont("Helvetica-Bold", 13)

    pdf.drawString(
        50,
        y,
        "RESUMO DE PAGAMENTOS"
    )

    y -= 25

    pdf.setFont("Helvetica", 11)

    pdf.drawString(
        50,
        y,
        f"Dinheiro: R$ {total_dinheiro:.2f}"
    )

    y -= 18

    pdf.drawString(
        50,
        y,
        f"Pix: R$ {total_pix:.2f}"
    )

    y -= 18

    pdf.drawString(
        50,
        y,
        f"Cartão: R$ {total_cartao:.2f}"
    )

    y -= 18

    pdf.drawString(
        50,
        y,
        f"Fiado: R$ {total_fiado:.2f}"
    )

    y -= 25

    pdf.setFont("Helvetica-Bold", 12)

    pdf.drawString(
        50,
        y,
        f"TOTAL VENDIDO: R$ {total_geral:.2f}"
    )

    y -= 40

    pdf.setFont("Helvetica-Bold", 13)

    pdf.drawString(
        50,
        y,
        "VENDAS DO DIA"
    )

    y -= 25

    pdf.setFont("Helvetica-Bold", 10)

    pdf.drawString(50, y, "ID")
    pdf.drawString(90, y, "CLIENTE")
    pdf.drawString(300, y, "PAGAMENTO")
    pdf.drawString(430, y, "TOTAL")

    y -= 15

    pdf.setFont("Helvetica", 10)

    for venda in vendas:

        if y < 70:
            pdf.showPage()
            y = 800

        pdf.drawString(
            50,
            y,
            str(venda.id)
        )

        pdf.drawString(
            90,
            y,
            venda.cliente.nome[:28]
        )

        pdf.drawString(
            300,
            y,
            venda.forma_pagamento or "-"
        )

        pdf.drawString(
            430,
            y,
            f"R$ {venda.total:.2f}"
        )

        y -= 18

    y -= 30

    pdf.drawString(
        50,
        y,
        "Sistema IvanTech ERP"
    )

    pdf.save()

    return response